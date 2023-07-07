import pandas as pd
from openpyxl import load_workbook,Workbook
from matplotlib import pyplot as plt


'''
（1）对Sheet1“停车情况记录表”中的“单价”列进行填充。（2分）
'''
book = load_workbook(r'原始数据.xlsx')
ws1 = book['Sheet1']
# 单价列 
for row in ws1.iter_rows(min_row=9,min_col=1,max_col=2):
    if row[1].value == '小汽车':
        ws1['C'+str(row[1].row)]=5
    elif row[1].value == '中客车':
        ws1['C' + str(row[1].row)] = 8
    else:
        ws1['C' + str(row[1].row)] = 10
book.save(r'原始数据.xlsx')


'''
（2）在Sheet1中，计算汽车在停车库中的停放时间。要求：* 计算方法为：“停放时间=出库时间-入库时间”。* 格式为：“小时：分钟：秒”。* 将结果保存在“停车情况记录表”中的“停放时间”列中。（2分）
（3）对“停车情况记录表”的停车费用进行计算。要求：* 根据Sheet1停放时间的长短计算停车费用，将计算结果填入到“停车情况记录表”的“应付金额”列中。
注意：* 停车按小时收费，对于不满一个小时的按照一个小时计费；
* 对于超过整点小时数十五分钟（包含十五分钟）的多累积一个小时。
* 例如1小时23分，将以2小时计费。（4分）
'''
df = pd.read_excel(r'原始数据.xlsx', sheet_name='Sheet1', header=7, usecols=range(2, 7))
# 停放时间列
# 先把时间列转换为datetime格式，才能计算时间差，得出的结果是timedelta格式
# timedelta数据类型,只能使用.days或.seconds提取时间再进行计算转换其他格式
df['入库时间'] = pd.to_datetime(df['入库时间'], format='%H:%M:%S')
df['出库时间'] = pd.to_datetime(df['出库时间'], format='%H:%M:%S')
df['停放时间'] = df['出库时间'] - df['入库时间']

# 应付金额列
def paypark(x):
    h = int(x['停放时间'].seconds / 3600) # 将timedelta数据格式转化为秒并提取小时部分
    m = int((x['停放时间'].seconds % 3600) / 60) # 提取分钟部分
    s = (x['停放时间'].seconds % 3600) % 60 # 提取秒部分
    if h == 0:
        pay = x['单价']
    elif m >= 15 and h > 0:
        pay = x['单价'] * (h + 1)
    else:
        pay = x['单价'] * h
    return pay
df['应付金额'] = df.apply(paypark,axis=1)

# 停放时间的格式化
def timeformat(x):
    h = int(x.seconds / 3600)
    m = int((x.seconds % 3600) / 60)
    s = (x.seconds % 3600) % 60
    return '{}小时:{}分钟:{}秒'.format(h, m, s)

df['停放时间'] = df['停放时间'].map(timeformat)

# print(df)
# print(df.info())


'''
（4）根据统计情况描述（I8，I9），补充完成统计结果（J8，J9）。（2分）
'''
df1 = pd.read_excel(r'原始数据.xlsx',sheet_name='Sheet1',usecols=[0,6],header=7)
# 停车费用大于等于40元的停车记录条数
df_q=df1.query('应付金额>=40').车牌号.count()
# 最高的停车费用
df_max = df1['应付金额'].max()
# 将数据写入指定单元格位置
book = load_workbook(r'原始数据.xlsx')
ws1 = book['Sheet1']
ws1['J8']=df_q
ws1['J9']=df_max
book.save(r'原始数据.xlsx')








